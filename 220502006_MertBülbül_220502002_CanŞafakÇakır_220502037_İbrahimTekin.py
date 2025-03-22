import sys
import mysql.connector
from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QMessageBox, QVBoxLayout, QWidget, QLineEdit, \
    QDialog, QFormLayout, QComboBox, QLabel, QSpacerItem, QSizePolicy, QGroupBox,QCheckBox
from mysql.connector import Error
from PyQt5.QtGui import QIcon, QPixmap
from PyQt5.QtCore import Qt



class DepartmentManager:
    def __init__(self):
        self.connection = None


    def connect_to_db(self):
        try:
            self.connection = mysql.connector.connect(
                host='127.0.0.1',
                database='dersprogrami',
                user='root',
                password='',
                port=3306
            )
            if self.connection.is_connected():
                return True
            else:
                return False
        except Error as e:
            print("Veritabanı bağlantısı hatası:", e)  # Detaylı hata mesajı
            return False

    def check_exists(self, table, column, value):
        if not self.connection or not self.connection.is_connected():
            self.connect_to_db()

        try:
            cursor = self.connection.cursor()
            query = f"SELECT COUNT(*) FROM {table} WHERE {column} = %s"
            cursor.execute(query, (value,))
            result = cursor.fetchone()[0]
            cursor.close()
            return result > 0
        except Error as e:
            print("Veritabanı sorgu hatası:", e)
            return False

    def add_classroom(self, classroom_name, capacity, type_):
        if not self.connection or not self.connection.is_connected():
            print("Bağlantı sağlanamadı, yeniden bağlanılıyor...")
            if not self.connect_to_db():
                print("Veritabanı bağlantısı sağlanamadı.")
                return False

        try:
            cursor = self.connection.cursor()
            query = "INSERT INTO classrooms (name, capacity, type) VALUES (%s, %s, %s)"
            values = (classroom_name, capacity, type_)
            cursor.execute(query, values)
            self.connection.commit()
            cursor.close()
            return True
        except Error as e:
            print("Veritabanı işlemi sırasında hata oluştu:", e)  # Detaylı hata mesajı
            return False

    def add_department(self, department_name, department_code):

        if not self.connection or not self.connection.is_connected():
            print("Bağlantı yok, yeni bağlantı açılıyor...")
            if not self.connect_to_db():
                print("Bağlantı sağlanamadı!")
                return False

        try:
            cursor = self.connection.cursor()
            query = "INSERT INTO departments (name, code) VALUES (%s, %s)"
            values = (department_name, department_code)
            cursor.execute(query, values)
            self.connection.commit()
            cursor.close()
            print("Bölüm başarıyla eklendi.")
            return True
        except Error as e:
            print("Veritabanı işlemi sırasında hata oluştu:", e)
            return False

    def add_course(self, name, code, weekly_hours, instructor_name, class_receiving, hangisinif, same_course_classes,
                   ikinci_sinif, ders_online_mi):
        if not self.connection or not self.connection.is_connected():
            print("Bağlantı sağlanamadı, yeniden bağlanılıyor...")
            if not self.connect_to_db():
                print("Veritabanı bağlantısı sağlanamadı.")
                return False

        try:
            cursor = self.connection.cursor()
            query = """
                INSERT INTO courses (name, code, weekly_hours, instructor_name, `DersiHangiBolumAliyor`, `HangiSinif`, `AyniDersiAlanBolum`, `HangiİkincilSinif`, `DersOnlineMi`)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """

            values = (
            name, code, weekly_hours, instructor_name, class_receiving, hangisinif, same_course_classes, ikinci_sinif,
            ders_online_mi)
            cursor.execute(query, values)
            self.connection.commit()
            cursor.close()
            return True
        except Error as e:
            print("Veritabanı işlemi sırasında hata oluştu:", e)
            return False

    def delete_classroom(self, classroom_name):
        if not self.connection or not self.connection.is_connected():
            if not self.connect_to_db():
                return False
        try:
            cursor = self.connection.cursor()
            query = "DELETE FROM classrooms WHERE name = %s"
            cursor.execute(query, (classroom_name,))
            self.connection.commit()
            cursor.close()
            return True
        except Error as e:
            print("Derslik silme hatası:", e)
            return False

    def delete_course(self, course_code):
        if not self.connection or not self.connection.is_connected():
            if not self.connect_to_db():
                return False
        try:
            cursor = self.connection.cursor()
            query = "DELETE FROM courses WHERE code = %s"
            cursor.execute(query, (course_code,))
            self.connection.commit()
            cursor.close()
            return True
        except Error as e:
            print("Ders silme hatası:", e)
            return False

    def add_student(self, name, student_no, department, class_):
        if not self.connection or not self.connection.is_connected():
            if not self.connect_to_db():
                print("Veritabanı bağlantısı sağlanamadı.")
                return False

        try:
            cursor = self.connection.cursor()
            query = "INSERT INTO students (İsimSoyisim, OgrenciNo, bolum, sinif) VALUES (%s, %s, %s, %s)"
            values = (name, student_no, department, class_)
            cursor.execute(query, values)
            self.connection.commit()
            cursor.close()
            return True
        except Error as e:
            print("Veritabanı işlemi sırasında hata oluştu:", e)
            return False


    def delete_department(self, department_code):
        if not self.connection or not self.connection.is_connected():
            if not self.connect_to_db():
                return False
        try:
            cursor = self.connection.cursor()
            query = "DELETE FROM departments WHERE code = %s"
            cursor.execute(query, (department_code,))
            self.connection.commit()
            cursor.close()
            return cursor.rowcount > 0
        except Error as e:
            print("Bölüm silme hatası:", e)
            return False

    def delete_student(self, student_no):
        if not self.connection or not self.connection.is_connected():
            if not self.connect_to_db():
                print("Veritabanı bağlantısı sağlanamadı.")
                return False

        try:
            cursor = self.connection.cursor()
            query = "DELETE FROM students WHERE OgrenciNo = %s"
            cursor.execute(query, (student_no,))
            self.connection.commit()
            cursor.close()
            return cursor.rowcount > 0
        except Error as e:
            print("Veritabanı işlemi sırasında hata oluştu:", e)
            return False

    def add_faculty_member(self, name):
        if not self.connection or not self.connection.is_connected():
            if not self.connect_to_db():
                print("Veritabanı bağlantısı sağlanamadı.")
                return False

        try:
            cursor = self.connection.cursor()
            query = "INSERT INTO facultymembers (İsimSoyisim) VALUES (%s)"
            cursor.execute(query, (name,))
            self.connection.commit()
            cursor.close()
            return True
        except Error as e:
            print("Veritabanı işlemi sırasında hata oluştu:", e)
            return False

    def delete_faculty_member(self, name):
        if not self.connection or not self.connection.is_connected():
            if not self.connect_to_db():
                print("Veritabanı bağlantısı sağlanamadı.")
                return False

        try:
            cursor = self.connection.cursor()
            query = "DELETE FROM facultymembers WHERE İsimSoyisim = %s"
            cursor.execute(query, (name,))
            self.connection.commit()
            cursor.close()
            return True
        except Error as e:
            print("Veritabanı işlemi sırasında hata oluştu:", e)
            return False

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("📘 Ders Programı Yönetim Sistemi")
        self.setGeometry(100, 100, 900, 1000)
        self.setStyleSheet("QMainWindow { background-color: #f0f2f5; }")
        self.setWindowIcon(QIcon("kostü.png"))

        self.department_manager = DepartmentManager()
        self.department_manager.connect_to_db()

        layout = QVBoxLayout()

        logo = QLabel()
        logo.setPixmap(QPixmap("kostü.png").scaled(100, 100, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        logo.setAlignment(Qt.AlignCenter)
        layout.addWidget(logo)

        title = QLabel("Ders Programı İşlemleri")
        title.setStyleSheet("font-size: 18px; font-weight: bold; margin: 10px;")
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        dept_group = QGroupBox("Bölüm İşlemleri")
        dept_layout = QVBoxLayout()
        dept_layout.addWidget(self.create_button("Bölüm Ekle", self.open_department_add_dialog))
        dept_layout.addWidget(self.create_button("Bölüm Sil", self.open_department_delete_dialog))
        dept_group.setLayout(dept_layout)
        layout.addWidget(dept_group)

        class_group = QGroupBox("Derslik İşlemleri")
        class_layout = QVBoxLayout()
        class_layout.addWidget(self.create_button("Derslik Ekle", self.open_classroom_add_dialog))
        class_layout.addWidget(self.create_button("Derslik Sil", self.open_classroom_delete_dialog))
        class_group.setLayout(class_layout)
        layout.addWidget(class_group)

        course_group = QGroupBox("Ders İşlemleri")
        course_layout = QVBoxLayout()
        course_layout.addWidget(self.create_button("Ders Ekle", self.open_course_add_dialog))
        course_layout.addWidget(self.create_button("Ders Sil", self.open_course_delete_dialog))
        course_group.setLayout(course_layout)
        layout.addWidget(course_group)

        stundents_group = QGroupBox("Öğrenci İşlemleri")
        student_layout = QVBoxLayout()
        student_layout.addWidget(self.create_button("Öğrenci Ekle", self.open_student_add_dialog))
        student_layout.addWidget(self.create_button("Öğrenci Sil", self.open_student_delete_dialog))
        stundents_group.setLayout(student_layout)
        layout.addWidget(stundents_group)

        faculty_group = QGroupBox("Öğretmen İşlemleri")
        faculty_layout = QVBoxLayout()
        faculty_layout.addWidget(self.create_button("Öğretmen Ekle", self.open_faculty_add_dialog))
        faculty_layout.addWidget(self.create_button("Öğretmen Sil", self.open_faculty_delete_dialog))
        faculty_group.setLayout(faculty_layout)
        layout.addWidget(faculty_group)

        schedule_group = QGroupBox("Ders Programı İşlemleri")
        schedule_layout = QVBoxLayout()

        schedule_layout.addWidget(self.create_button("Programı Oluştur", self.ders_programi_olustur))

        schedule_group.setLayout(schedule_layout)
        layout.addWidget(schedule_group)

        self.result_label = QLabel("")
        self.result_label.setAlignment(Qt.AlignCenter)
        self.result_label.setStyleSheet("font-size: 13px; color: #333; margin-top: 20px;")
        layout.addWidget(self.result_label)


        widget = QWidget()
        widget.setLayout(layout)
        self.setCentralWidget(widget)

        # Diğer İşlemler
        layout.addWidget(self.create_button("Bağlantı Testi", self.test_connection))
        layout.addSpacerItem(QSpacerItem(20, 40, QSizePolicy.Minimum, QSizePolicy.Expanding))

        container = QWidget()
        container.setLayout(layout)
        self.setCentralWidget(container)

    def create_button(self, text, handler):
        button = QPushButton(text)
        button.clicked.connect(handler)
        button.setStyleSheet(
            """
            QPushButton {
                background-color: #27ae60; 
                color: white; 
                padding: 10px; 
                border-radius: 6px; 
                font-weight: bold; 
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #219150;
            }
            """
        )
        return button

    def test_connection(self):
        if self.department_manager.connect_to_db():
            QMessageBox.information(self, "Bağlantı Durumu", "✅ Veritabanı bağlantısı başarılı!")
        else:
            QMessageBox.critical(self, "Bağlantı Hatası", "❌ Veritabanı bağlantısı sağlanamadı!")

    def open_department_add_dialog(self):
        dialog = DepartmentAddDialog()
        dialog.setStyleSheet(self.dialog_style())
        dialog.exec_()

    def open_department_delete_dialog(self):
        dialog = DepartmentDeleteDialog()
        dialog.setStyleSheet(self.dialog_style())
        dialog.exec_()

    def open_classroom_add_dialog(self):
        dialog = ClassroomAddDialog()
        dialog.setStyleSheet(self.dialog_style())
        dialog.exec_()

    def open_classroom_delete_dialog(self):
        dialog = ClassroomDeleteDialog()
        dialog.setStyleSheet(self.dialog_style())
        dialog.exec_()

    def open_course_add_dialog(self):
        dialog = CourseAddDialog()
        dialog.setStyleSheet(self.dialog_style())
        dialog.exec_()

    def open_course_delete_dialog(self):
        dialog = CourseDeleteDialog()
        dialog.setStyleSheet(self.dialog_style())
        dialog.exec_()

    def open_student_add_dialog(self):
        dialog = StudentAddDialog()
        dialog.setStyleSheet(self.dialog_style())
        dialog.exec_()

    def open_student_delete_dialog(self):
        dialog = StudentDeleteDialog()
        dialog.setStyleSheet(self.dialog_style())
        dialog.exec_()

    def open_faculty_add_dialog(self):
        dialog = FacultyAddDialog()
        dialog.setStyleSheet(self.dialog_style())
        dialog.exec_()

    def open_faculty_delete_dialog(self):
        dialog = FacultyDeleteDialog()
        dialog.setStyleSheet(self.dialog_style())
        dialog.exec_()

    def ders_programi_olustur(self):
        if not hasattr(self, "department_manager"):
            print("❌ DepartmentManager tanımlı değil!")
            QMessageBox.critical(self, "Hata", "DepartmentManager başlatılmadı!")
            return

        if not self.department_manager.connection or not self.department_manager.connection.is_connected():
            print("⚠️ Veritabanı bağlantısı yok, bağlanılıyor...")
            if not self.department_manager.connect_to_db():
                print("❌ Veritabanına bağlanılamadı!")
                QMessageBox.critical(self, "Hata", "Veritabanına bağlanılamadı!")
                return

        try:
            bolumler = [("BLM", "bilgisayar_muhendisligi_programi.xlsx"), ("YZM", "yazilim_muhendisligi_programi.xlsx")]

            for bolum_adi, excel_dosya in bolumler:
                print(f"📌 {bolum_adi} ders programı oluşturma işlemi başladı...")
                program_olusturucu = DersProgramiOlusturucu(self.department_manager.connection, bolum_adi, excel_dosya)

                print(f"📌 {bolum_adi} verileri yükleniyor...")
                program_olusturucu.verileri_yukle()

                print(f"📌 {bolum_adi} dersleri işleniyor...")
                program_olusturucu.ders_programini_olustur()

                print(f"📌 {bolum_adi} Excel dosyasına yazılıyor...")
                program_olusturucu.excel_yaz()

            print("✅ Tüm ders programları başarıyla oluşturuldu!")
            QMessageBox.information(self, "Başarı", "Tüm ders programları başarıyla oluşturuldu ve Excel'e yazıldı!")
        except Exception as e:
            print("❌ Hata oluştu:", e)
            QMessageBox.critical(self, "Hata", f"Bir hata oluştu: {e}")

    def dialog_style(self):
        return """
            QDialog {
                background-color: #ffffff;
            }
            QLabel {
                font-size: 14px;
                color: #2c3e50;
            }
            QLineEdit, QComboBox {
                padding: 6px;
                font-size: 13px;
            }
            QPushButton {
                background-color: #27ae60;
                color: white;
                padding: 8px;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #219150;
            }
        """


class DepartmentAddDialog(QDialog):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Bölüm Ekle")
        self.setGeometry(200, 200, 300, 200)

        layout = QFormLayout()

        self.department_name_input = QLineEdit(self)
        self.department_code_input = QLineEdit(self)

        self.department_name_input.setPlaceholderText("Bölüm Adı")
        self.department_code_input.setPlaceholderText("Bölüm Kodu")

        self.add_button = QPushButton("Ekle", self)
        self.add_button.clicked.connect(self.add_department)

        layout.addRow("Bölüm Adı:", self.department_name_input)
        layout.addRow("Bölüm Kodu:", self.department_code_input)
        layout.addWidget(self.add_button)

        self.setLayout(layout)

        self.department_manager = DepartmentManager()

    def add_department(self):
        department_name = self.department_name_input.text()
        department_code = self.department_code_input.text()

        if department_name and department_code:
            if self.department_manager.add_department(department_name, department_code):
                QMessageBox.information(self, "Başarı", "Bölüm başarıyla eklendi.")
                self.accept()
            else:
                QMessageBox.warning(self, "Hata", "Bölüm eklenirken bir hata oluştu.")
        else:
            QMessageBox.warning(self, "Eksik Veri", "Bölüm adı ve kodu boş olamaz!")



class ClassroomAddDialog(QDialog):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Derslik Ekle")
        self.setGeometry(200, 200, 300, 250)

        layout = QFormLayout()

        self.classroom_name_input = QLineEdit(self)
        self.capacity_input = QLineEdit(self)
        self.type_input = QComboBox(self)
        self.type_input.addItem("NORMAL")
        self.type_input.addItem("LAB")

        self.classroom_name_input.setPlaceholderText("Derslik Adı")
        self.capacity_input.setPlaceholderText("Kapasite")

        self.add_button = QPushButton("Ekle", self)
        self.add_button.clicked.connect(self.add_classroom)

        layout.addRow("Derslik Adı:", self.classroom_name_input)
        layout.addRow("Kapasite:", self.capacity_input)
        layout.addRow("Türü:", self.type_input)
        layout.addWidget(self.add_button)

        self.setLayout(layout)

        self.department_manager = DepartmentManager()

    def add_classroom(self):
        classroom_name = self.classroom_name_input.text()
        capacity = self.capacity_input.text()
        type_ = self.type_input.currentText()

        if classroom_name and capacity and type_:
            if self.department_manager.add_classroom(classroom_name, capacity, type_):
                QMessageBox.information(self, "Başarı", "Derslik başarıyla eklendi.")
                self.accept()
            else:
                QMessageBox.warning(self, "Hata", "Veritabanı bağlantısı sağlanamadı veya başka bir hata oluştu.")
        else:
            QMessageBox.warning(self, "Eksik Veri", "Tüm alanlar doldurulmalıdır!")


class CourseAddDialog(QDialog):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Ders Ekle")
        self.setGeometry(200, 200, 300, 300)

        layout = QFormLayout()

        self.course_name_input = QLineEdit(self)
        self.course_code_input = QLineEdit(self)
        self.weekly_hours_input = QLineEdit(self)
        self.instructor_id_input = QLineEdit(self)
        self.class_receiving_course_input = QLineEdit(self)
        self.hangisinif_input = QLineEdit(self)
        self.same_course_classes_input = QLineEdit(self)
        self.ikinci_sinif_input = QLineEdit(self)

        self.course_name_input.setPlaceholderText("Ders Adı")
        self.course_code_input.setPlaceholderText("Ders Kodu")
        self.weekly_hours_input.setPlaceholderText("Haftalık Saat")
        self.instructor_id_input.setPlaceholderText("Eğitmen Adı")
        self.class_receiving_course_input.setPlaceholderText("Dersi Alan Bölüm")
        self.hangisinif_input.setPlaceholderText("Kaçıncı Sınıf")
        self.same_course_classes_input.setPlaceholderText("Aynı Dersi Alan Sınıflar (Opsiyonel)")
        self.ikinci_sinif_input.setPlaceholderText("Kaçıncı Sınıf")

        self.ders_online_checkbox = QCheckBox("Ders Online Mi?", self)

        self.add_button = QPushButton("Ekle", self)
        self.add_button.clicked.connect(self.add_course)

        layout.addRow("Ders Adı:", self.course_name_input)
        layout.addRow("Ders Kodu:", self.course_code_input)
        layout.addRow("Haftalık Saat:", self.weekly_hours_input)
        layout.addRow("Eğitmen Adı:", self.instructor_id_input)
        layout.addRow("Dersi Alan Bölüm:", self.class_receiving_course_input)
        layout.addRow("Kaçıncı Sınıf", self.hangisinif_input)
        layout.addRow("Aynı Dersi Alan Bölüm:", self.same_course_classes_input)
        layout.addRow("Kaçıncı Sınıf", self.ikinci_sinif_input)
        layout.addRow(self.ders_online_checkbox)
        layout.addWidget(self.add_button)

        self.setLayout(layout)

        self.department_manager = DepartmentManager()

    def add_course(self):
        name = self.course_name_input.text()
        code = self.course_code_input.text()
        weekly_hours = self.weekly_hours_input.text()
        instructor_name = self.instructor_id_input.text()
        class_receiving = self.class_receiving_course_input.text()
        hangisinif = self.hangisinif_input.text()
        same_course_classes = self.same_course_classes_input.text() or None
        ikinci_sinif = self.ikinci_sinif_input.text() or None
        ders_online_mi = self.ders_online_checkbox.isChecked()

        if name and code and weekly_hours and instructor_name and class_receiving and hangisinif:
            cursor = self.department_manager.connection.cursor()
            cursor.execute("SELECT COUNT(*) FROM courses WHERE code = %s", (code,))
            if cursor.fetchone()[0] > 0:
                cursor.close()
                QMessageBox.warning(self, "Uyarı", "Bu ders kodu zaten kayıtlı!")
                return
            cursor.close()

            if self.department_manager.add_course(name, code, weekly_hours, instructor_name, class_receiving,
                                                  hangisinif, same_course_classes, ikinci_sinif, ders_online_mi):
                QMessageBox.information(self, "Başarı", "Ders başarıyla eklendi.")
                self.accept()
            else:
                QMessageBox.warning(self, "Hata", "Ders eklenirken bir hata oluştu.")
        else:
            QMessageBox.warning(self, "Eksik Veri", "Zorunlu alanlar boş olamaz!")


class DepartmentDeleteDialog(QDialog):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Bölüm Sil")
        self.setGeometry(200, 200, 300, 150)

        layout = QFormLayout()

        self.department_code_input = QLineEdit(self)
        self.department_code_input.setPlaceholderText("Bölüm Kodu")

        self.delete_button = QPushButton("Sil", self)
        self.delete_button.clicked.connect(self.delete_department)

        layout.addRow("Bölüm Kodu:", self.department_code_input)
        layout.addWidget(self.delete_button)

        self.setLayout(layout)

        self.department_manager = DepartmentManager()

    def delete_department(self):
        department_code = self.department_code_input.text()
        if department_code:
            if self.department_manager.delete_department(department_code):
                QMessageBox.information(self, "Başarı", "Bölüm başarıyla silindi.")
                self.accept()
            else:
                QMessageBox.warning(self, "Hata", "Silme işlemi başarısız.")
        else:
            QMessageBox.warning(self, "Eksik Veri", "Bölüm kodu girilmelidir!")

class ClassroomDeleteDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Derslik Sil")
        self.setGeometry(200, 200, 300, 150)

        layout = QFormLayout()
        self.classroom_name_input = QLineEdit(self)
        self.classroom_name_input.setPlaceholderText("Derslik Adı")

        self.delete_button = QPushButton("Sil", self)
        self.delete_button.clicked.connect(self.delete_classroom)

        layout.addRow("Derslik Adı:", self.classroom_name_input)
        layout.addWidget(self.delete_button)
        self.setLayout(layout)

        self.department_manager = DepartmentManager()

    def delete_classroom(self):
        classroom_name = self.classroom_name_input.text()
        if classroom_name:
            reply = QMessageBox.question(
                self, "Emin misiniz?",
                f"{classroom_name} adlı dersliği silmek istiyor musunuz?",
                QMessageBox.Yes | QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                if self.department_manager.delete_classroom(classroom_name):
                    QMessageBox.information(self, "Başarı", "Derslik silindi.")
                    self.accept()
                else:
                    QMessageBox.warning(self, "Hata", "Silme işlemi başarısız.")
        else:
            QMessageBox.warning(self, "Eksik Veri", "Derslik adı girilmelidir!")

class CourseDeleteDialog(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Ders Sil")
        self.setGeometry(200, 200, 300, 150)

        layout = QFormLayout()
        self.course_code_input = QLineEdit(self)
        self.course_code_input.setPlaceholderText("Ders Kodu")

        self.delete_button = QPushButton("Sil", self)
        self.delete_button.clicked.connect(self.delete_course)

        layout.addRow("Ders Kodu:", self.course_code_input)
        layout.addWidget(self.delete_button)
        self.setLayout(layout)

        self.department_manager = DepartmentManager()

    def delete_course(self):
        course_code = self.course_code_input.text()
        if course_code:
            reply = QMessageBox.question(
                self, "Emin misiniz?",
                f"{course_code} kodlu dersi silmek istiyor musunuz?",
                QMessageBox.Yes | QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                if self.department_manager.delete_course(course_code):
                    QMessageBox.information(self, "Başarı", "Ders silindi.")
                    self.accept()
                else:
                    QMessageBox.warning(self, "Hata", "Silme işlemi başarısız.")
        else:
            QMessageBox.warning(self, "Eksik Veri", "Ders kodu girilmelidir!")

class ClassAddDialog(QDialog):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Sınıf Ekle")
        self.setGeometry(150, 150, 300, 250)

        layout = QFormLayout()

        self.id_input = QLineEdit()
        layout.addRow("ID:", self.id_input)

        self.department_input = QLineEdit()
        layout.addRow("Bölüm:", self.department_input)

        self.class_name_input = QLineEdit()
        layout.addRow("Sınıf Adı:", self.class_name_input)

        self.capacity_input = QLineEdit()
        layout.addRow("Kapasite:", self.capacity_input)

        self.save_button = QPushButton("Kaydet")
        self.save_button.clicked.connect(self.save_class)
        layout.addRow(self.save_button)

        self.setLayout(layout)

    def save_class(self):
        id_value = self.id_input.text()
        department_value = self.department_input.text()
        class_name_value = self.class_name_input.text()
        capacity_value = self.capacity_input.text()

        if not (id_value and department_value and class_name_value and capacity_value):
            QMessageBox.warning(self, "Hata", "Lütfen tüm alanları doldurun!")
            return

        try:
            conn = mysql.connector.connect(host='localhost', user='root', password='', database='dersprogrami')
            cursor = conn.cursor()
            query = "INSERT INTO classes (id, bolum, sinif, kapasite) VALUES (%s, %s, %s, %s)"
            cursor.execute(query, (id_value, department_value, class_name_value, capacity_value))
            conn.commit()
            conn.close()

            QMessageBox.information(self, "Başarılı", "Sınıf başarıyla eklendi!")
            self.close()

        except mysql.connector.Error as e:
            QMessageBox.critical(self, "Veritabanı Hatası", f"Hata: {str(e)}")

class StudentAddDialog(QDialog):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Öğrenci Ekle")
        self.setGeometry(200, 200, 300, 300)

        layout = QFormLayout()

        self.name_input = QLineEdit(self)
        self.student_no_input = QLineEdit(self)
        self.department_input = QLineEdit(self)
        self.class_input = QLineEdit(self)

        self.name_input.setPlaceholderText("İsim Soyisim")
        self.student_no_input.setPlaceholderText("Öğrenci No")
        self.department_input.setPlaceholderText("Bölüm")
        self.class_input.setPlaceholderText("Sınıf")

        self.add_button = QPushButton("Öğrenci Ekle", self)
        self.add_button.clicked.connect(self.add_student)

        layout.addRow("İsim Soyisim:", self.name_input)
        layout.addRow("Öğrenci No:", self.student_no_input)
        layout.addRow("Bölüm:", self.department_input)
        layout.addRow("Sınıf:", self.class_input)
        layout.addWidget(self.add_button)

        self.setLayout(layout)

        self.student_manager = DepartmentManager()

    def add_student(self):
        name = self.name_input.text()
        student_no = self.student_no_input.text()
        department = self.department_input.text()
        class_ = self.class_input.text()

        if name and student_no and department and class_:
            if self.student_manager.check_exists("students", "OgrenciNo", student_no):
                QMessageBox.warning(self, "Uyarı", "Bu öğrenci numarası zaten kayıtlı!")
                return

            if self.student_manager.add_student(name, student_no, department, class_):
                QMessageBox.information(self, "Başarı", "Öğrenci başarıyla eklendi.")
                self.accept()
            else:
                QMessageBox.warning(self, "Hata", "Veritabanı işlemi sırasında hata oluştu.")
        else:
            QMessageBox.warning(self, "Eksik Veri", "Tüm alanlar doldurulmalıdır!")


class StudentDeleteDialog(QDialog):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Öğrenci Sil")
        self.setGeometry(200, 200, 300, 150)

        layout = QFormLayout()

        self.student_no_input = QLineEdit(self)
        self.student_no_input.setPlaceholderText("Öğrenci No")

        self.delete_button = QPushButton("Öğrenci Sil", self)
        self.delete_button.clicked.connect(self.delete_student)

        layout.addRow("Öğrenci No:", self.student_no_input)
        layout.addWidget(self.delete_button)

        self.setLayout(layout)

        self.student_manager = DepartmentManager()

    def delete_student(self):
        student_no = self.student_no_input.text()
        if student_no:
            reply = QMessageBox.question(
                self, "Emin misiniz?",
                f"{student_no} numaralı öğrenciyi silmek istiyor musunuz?",
                QMessageBox.Yes | QMessageBox.No
            )
            if reply == QMessageBox.Yes:
                if self.department_manager.delete_student(student_no):
                    QMessageBox.information(self, "Başarı", "Öğrenci başarıyla silindi.")
                    self.accept()
                else:
                    QMessageBox.warning(self, "Hata", "Öğrenci bulunamadı veya silme işlemi başarısız.")
        else:
            QMessageBox.warning(self, "Eksik Veri", "Öğrenci numarası girilmelidir!")

class FacultyAddDialog(QDialog):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Öğretmen Ekle")
        self.setGeometry(200, 200, 300, 300)

        layout = QFormLayout()

        self.name_input = QLineEdit(self)
        self.name_input.setPlaceholderText("İsim Soyisim")

        self.add_button = QPushButton("Öğretmen Ekle", self)
        self.add_button.clicked.connect(self.add_faculty_member)

        layout.addRow("İsim Soyisim:", self.name_input)
        layout.addWidget(self.add_button)

        self.setLayout(layout)

        self.faculty_manager = DepartmentManager()

    def add_faculty_member(self):
        name = self.name_input.text()

        if name:
            if self.faculty_manager.add_faculty_member(name):
                QMessageBox.information(self, "Başarı", "Öğretmen başarıyla eklendi.")
                self.accept()
            else:
                QMessageBox.warning(self, "Hata", "Veritabanı işlemi sırasında hata oluştu.")
        else:
            QMessageBox.warning(self, "Eksik Veri", "Tüm alanlar doldurulmalıdır!")


class FacultyDeleteDialog(QDialog):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Öğretmen Sil")
        self.setGeometry(200, 200, 300, 150)

        layout = QFormLayout()

        self.name_input = QLineEdit(self)
        self.name_input.setPlaceholderText("İsim Soyisim")

        self.delete_button = QPushButton("Öğretmen Sil", self)
        self.delete_button.clicked.connect(self.delete_faculty_member)

        layout.addRow("İsim Soyisim:", self.name_input)
        layout.addWidget(self.delete_button)

        self.setLayout(layout)

        self.faculty_manager = DepartmentManager()

    def delete_faculty_member(self):
        name = self.name_input.text()

        if name:
            if self.faculty_manager.delete_faculty_member(name):
                QMessageBox.information(self, "Başarı", "Öğretmen başarıyla silindi.")
                self.accept()
            else:
                QMessageBox.warning(self, "Hata", "Öğretmen bulunamadı veya silme işlemi başarısız.")
        else:
            QMessageBox.warning(self, "Eksik Veri", "Öğretmen adı girilmelidir!")


# Güncellenmiş DersProgramiOlusturucu
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
import pandas as pd
import random

class DersProgramiOlusturucu:
    def __init__(self, connection, bolum_adi, excel_dosya):
        self.connection = connection
        self.cursor = self.connection.cursor(dictionary=True)
        self.excel_dosya = excel_dosya
        self.bolum_adi = bolum_adi

        self.gunler = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]
        self.normal_saatler = ["09:00-10:00", "10:00-11:00", "11:00-12:00", "13:00-14:00", "14:00-15:00", "15:00-16:00", "16:00-17:00"]
        self.online_saatler = ["15:00-16:00","16:00-17:00"]
        self.tum_saatler = self.normal_saatler + self.online_saatler

        self.program = {
            gun: {saat: {"1. Sınıf": None, "2. Sınıf": None, "3. Sınıf": None, "4. Sınıf": None} for saat in self.tum_saatler}
            for gun in self.gunler
        }

    def verileri_yukle(self):
        self.cursor.execute("""
            SELECT * FROM courses 
            WHERE DersiHangiBolumAliyor = %s OR FIND_IN_SET(%s, AyniDersiAlanBolum)
        """, (self.bolum_adi, self.bolum_adi))
        self.dersler = self.cursor.fetchall()

        self.cursor.execute("SELECT * FROM students WHERE bolum = %s", (self.bolum_adi,))
        self.ogrenciler = self.cursor.fetchall()

        self.cursor.execute("SELECT * FROM classrooms")
        self.siniflar = self.cursor.fetchall()

        self.cursor.execute("SELECT bolum, sinif, COUNT(*) as ogr_sayisi FROM students GROUP BY bolum, sinif")
        self.ogrenci_sayilari = {
            (row['bolum'], int(row['sinif'])): row['ogr_sayisi'] for row in self.cursor.fetchall()
        }

    def slot_musait_mi(self, gun, saat, sinif_key, hoca):
        if self.program[gun][saat][sinif_key] is not None:
            return False
        for mevcut in self.program[gun][saat].values():
            if mevcut and hoca in mevcut['ders']:
                return False
        return True

    def uygun_sinif_bul(self, kapasite_ihtiyaci, ders_tipi, gun, saat):
        kullanilan_siniflar = set()
        for v in self.program[gun][saat].values():
            if isinstance(v, str) and ' - ' in v:
                try:
                    kullanilan_siniflar.add(v.rsplit(' - ', 1)[-1])
                except Exception:
                    continue

        uygunlar = [
            s for s in self.siniflar
            if s.get("type") == ders_tipi
               and int(s.get("capacity", 0)) >= kapasite_ihtiyaci
               and s.get("name") not in kullanilan_siniflar
        ]
        random.shuffle(uygunlar)
        return uygunlar[0].get("name") if uygunlar else None

    def ders_programini_olustur(self):
        tum_dersler = self.dersler
        gun_sayaci = 0

        for ders in tum_dersler:
            if not ders.get("HangiSinif"):
                continue

            saat_ihtiyaci = int(ders['weekly_hours'])
            hoca = ders['instructor_name']
            is_online = ders.get('DersOnlineMi')
            saat_araligi = self.online_saatler if is_online else self.normal_saatler

            bolumler = [ders['DersiHangiBolumAliyor']]
            if ders.get('AyniDersiAlanBolum'):
                bolumler.extend((ders.get('AyniDersiAlanBolum') or '').split(','))

            if self.bolum_adi not in bolumler:
                continue

            sinif_key = f"{ders['HangiSinif']}. Sınıf"
            ogrenci_sayisi = sum(1 for o in self.ogrenciler if str(o.get("sinif")) == str(ders["HangiSinif"]))
            sinif_turu = "LAB" if "LAB" in (ders.get("name") or '').upper() else "NORMAL"

            yerlestirildi = False
            while saat_ihtiyaci > 0:
                gun = self.gunler[gun_sayaci % len(self.gunler)]
                gun_sayaci += 1

                for saat in saat_araligi:
                    if isinstance(self.program[gun][saat][sinif_key], str) and self.program[gun][saat][sinif_key]:
                        continue

                    if is_online:
                        sinif_adi = "ONLINE"
                    else:
                        sinif_adi = self.uygun_sinif_bul(ogrenci_sayisi, sinif_turu, gun, saat)
                        if not sinif_adi:
                            continue

                    self.program[gun][saat][sinif_key] = f"{ders['name']} ({hoca}) - {sinif_adi}"
                    saat_ihtiyaci -= 1
                    if saat_ihtiyaci == 0:
                        yerlestirildi = True
                        break
                if yerlestirildi:
                    break

    def yaz_bolum_adi(self, ws, hucre, text):
        for merge in ws.merged_cells.ranges:
            if hucre in merge:
                r, c = merge.min_row, merge.min_col
                ws.cell(row=r, column=c).value = text
                return
        ws[hucre] = text

    def excel_yaz(self):
        try:
            template_path = "BOŞ BÜTÜNLEŞİK ŞABLON - yazlab.xlsx"
            wb = load_workbook(template_path)
            ws = wb.active

            gun_satirlari = {
                "Pazartesi": 4,
                "Salı": 14,
                "Çarşamba": 23,
                "Perşembe": 28,
                "Cuma": 36
            }

            sinif_kolonlari = {
                "1. Sınıf": "C",
                "2. Sınıf": "D",
                "3. Sınıf": "E",
                "4. Sınıf": "F"
            }

            for gun in self.gunler:
                bas_satir = gun_satirlari[gun]
                for i, saat in enumerate(self.normal_saatler):
                    satir = bas_satir + i
                    for sinif, kolon in sinif_kolonlari.items():
                        icerik = self.program[gun][saat][sinif]
                        if icerik:
                            ws[f"{kolon}{satir}"] = icerik

            self.yaz_bolum_adi(ws, "D1", self.bolum_adi)

            dosya_adi = f"{self.bolum_adi.lower()}_ders_programi.xlsx"
            kayit_yolu = rf"C:\\Users\\ibrah\\PycharmProjects\\PythonProject11\\{dosya_adi}"
            wb.save(kayit_yolu)
            print(f"✅ Excel dosyası kaydedildi: {kayit_yolu}")

        except Exception as e:
            print("❌ Excel yazım hatası:", e)



def log_uncaught_exceptions(exctype, value, traceback):
    print("Beklenmeyen hata oluştu:", value)
    sys.__excepthook__(exctype, value, traceback)

sys.excepthook = log_uncaught_exceptions

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())